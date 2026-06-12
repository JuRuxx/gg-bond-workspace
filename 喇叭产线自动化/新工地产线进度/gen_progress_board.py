#!/usr/bin/env python3
"""
天勃项目进度跟踪表 → 可视化看板图
读取 xlsx 总表，输出 PNG 看板

用法:
    python3 gen_progress_board.py
    python3 gen_progress_board.py --sort due      # 按交付日期排序
    python3 gen_progress_board.py --title "天勃新厂项目"
"""

import os
import sys
import re
from datetime import datetime, timedelta
import openpyxl

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.font_manager import FontProperties, findfont, FontManager
import numpy as np

# ── 字体（复用看板生成器逻辑）─────────────────────────
CHINESE_FONTS = [
    'PingFang SC', 'PingFang', 'Heiti SC', 'Heiti TC',
    'STHeiti', 'Microsoft YaHei', 'SimHei',
    'WenQuanYi Micro Hei', 'Noto Sans CJK SC', 'Noto Sans SC',
    'Apple LiGothic', 'LiHei Pro',
]

def find_chinese_font():
    fm = FontManager()
    available = {f.name for f in fm.ttflist}
    for font in CHINESE_FONTS:
        if font in available:
            try:
                fp = FontProperties(family=font)
                findfont(fp, fallback_to_default=False)
                return font
            except Exception:
                continue
    for f in fm.ttflist:
        if any(c in f.name for c in ['Hei', 'Song', 'Ming', 'Fang', 'Kai', 'Noto', 'CJK', 'PingFang', 'YaHei']):
            return f.name
    return None


def excel_serial_to_date(serial):
    if isinstance(serial, (int, float)) and serial > 40000:
        return (datetime(1899, 12, 30) + timedelta(days=int(serial))).strftime('%Y-%m-%d')
    return None


# ── 状态 → 进度百分比 映射 ───────────────────────
# 生产流程阶段：结构出图 → 采购 → 机加件 → 装配 → 电气/软件 → 完成
# 总共约5步，每步20%

def stage_to_progress(item):
    """
    根据各字段内容推断完成百分比
    总分100%，拆分：
      - 结构出图: 20%
      - 采购: 25%
      - 机加件: 25%
      - 装配调试: 20%
      - 电气软件: 10%
    """
    struct = (item.get('结构') or '').strip()
    purchase = (item.get('采购') or '').strip()
    machining = (item.get('机加件') or '').strip()
    assembly = (item.get('装配（硬件+电控）') or '').strip()
    electric = (item.get('电气') or '').strip()
    software = (item.get('软件') or '').strip()
    remarks = (item.get('备注') or '').strip()

    # ── 阶段1: 结构（20%）──
    done_struct = 0
    if struct:
        if '己组装好' in struct or '已完成' in struct or '组装完成' in struct:
            done_struct = 20
        elif '己出图' in struct or '出图' in struct:
            done_struct = 20
        elif '组装中' in struct or '调试中' in struct:
            done_struct = 20
        elif '只报价暂不生产' in struct or '不生产' in struct:
            # 此类项目不参与进度计算
            return 0, '暂不进行'

    # ── 阶段2: 采购（25%）──
    done_purchase = 0
    if purchase:
        if purchase == '/' or purchase == '-':
            done_purchase = 25  # 不适用（如充磁机已组装好）
        elif '己到' in purchase or '已交' in purchase or '交进' in purchase:
            done_purchase = 25
        elif re.match(r'\d{4}-\d{2}-\d{2}', purchase):
            done_purchase = 25  # 有明确日期说明已采购
        elif '采购未回' in purchase or '报价中' in purchase or '未回' in purchase:
            done_purchase = 8
        elif purchase.strip() in ('', ' '):
            done_purchase = 0
        else:
            done_purchase = 20  # 其他状态（部分完成）

    # ── 阶段3: 机加件（25%）──
    done_machining = 0
    if machining:
        if machining == '/' or machining == '-':
            done_machining = 25
        elif '库存有料' in machining or '己到' in machining or '已到' in machining:
            done_machining = 25  # 有库存/已到料 = 完成
        elif re.match(r'\d{4}-\d{2}-\d{2}', machining):
            done_machining = 25
        elif '加工中' in machining or '己下料' in machining or '待加工' in machining:
            done_machining = 15  # 加工进行中
        elif '发出去喷砂' in machining or '己发' in machining:
            done_machining = 20  # 外协处理中
        elif machining.strip() in ('', ' '):
            done_machining = 0
        else:
            done_machining = 10

    # ── 阶段4: 装配（20%）──
    done_assembly = 0
    if assembly:
        if '调试' in assembly or '组装' in assembly:
            done_assembly = 15
        elif re.match(r'\d{4}-\d{2}-\d{2}', assembly):
            done_assembly = 20
        else:
            done_assembly = 10
    # 检查备注里有组装的线索
    if not assembly and remarks:
        if '己组装好' in struct or '组装完成' in remarks:
            done_assembly = 20
        elif '组装' in remarks or '接电气路' in remarks:
            done_assembly = 20

    # ── 阶段5: 电气/软件（10%）──
    done_electric = 0
    if electric:
        if '己出图' in electric or '出图' in electric:
            done_electric = 8
        elif '没有出图纸' in electric:
            done_electric = 2
        else:
            done_electric = 5
    if software:
        done_electric = max(done_electric, 5)

    # 特殊：已组装好的设备
    if '己组装好' in struct or '6台己组装好' in struct:
        return 100, '已完成'

    total = done_struct + done_purchase + done_machining + done_assembly + done_electric
    # 输出状态描述
    if total >= 100:
        status = '已完成'
    elif total >= 80:
        status = '装配调试中'
    elif total >= 55:
        status = '机加/装配中'
    elif total >= 30:
        status = '采购/机加中'
    elif total >= 10:
        status = '出图完成'
    else:
        status = '规划中'

    return min(total, 100), status


def parse_item_priority(item, progress, status):
    """
    根据剩余难度推断优先级
    - P0: 进度<30%（严重滞后）OR 临近交付但采购还没回
    - P1: 进度30-70%
    - P2: 进度>70% 或 已完成
    """
    name = item.get('设备名称', '').strip()
    purchase = (item.get('采购') or '').strip()
    due_str = item.get('出图时间') or ''

    if '只报价暂不生产' in (item.get('结构') or ''):
        return 'P2'

    if progress >= 100:
        return 'P2'

    # 采购没回+出图时间早于现在的 -> P0
    if '采购未回' in purchase or '未回' in purchase:
        return 'P0'

    if progress < 30:
        return 'P0'
    elif progress < 70:
        return 'P1'
    else:
        return 'P2'


def determine_due(item):
    """找出最晚的日期作为交付日期"""
    due = None
    for field in ['出图时间', '采购', '机加件', '装配（硬件+电控）']:
        val = item.get(field, '')
        if isinstance(val, (int, float)) and val > 45000:
            d = excel_serial_to_date(val)
            if d and (due is None or d > due):
                due = d
        elif isinstance(val, str):
            m = re.search(r'(\d+)/(\d+)', val)
            if m:
                d = f'2026-{int(m.group(1)):02d}-{int(m.group(2)):02d}'
                if due is None or d > due:
                    due = d
    return due or ''


# ── 配色方案（与看板生成器一致）───────────────────
BG_PAGE = '#F5F6F8'
BG_HEADER = '#2C3E50'
TEXT_HEADER = '#FFFFFF'
TEXT_TITLE = '#2C3E50'
TEXT_SUBTITLE = '#7B8A8B'

ROW_EVEN = '#FFFFFF'
ROW_ODD = '#F8F9FB'
ROW_BORDER = '#E8ECF0'

PRIORITY_COLORS = {
    'P0': {'bg': '#FDECEA', 'fg': '#C0392B', 'tag': '#E74C3C'},
    'P1': {'bg': '#FEF5E7', 'fg': '#D68910', 'tag': '#F39C12'},
    'P2': {'bg': '#EBF5FB', 'fg': '#1A5276', 'tag': '#2980B9'},
}

PROGRESS_COLORS = {
    'high': '#27AE60',
    'mid': '#2E86C1',
    'low': '#95A5A6',
}
PROGRESS_BG = '#E8ECF0'


def load_xlsx_data(filepath):
    """从xlsx总表加载设备项目列表"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['总表']

    items = []
    headers = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers = [str(v) if v else '' for v in row]

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
        vals = list(row)
        seq = vals[0]
        name = str(vals[1] or '').strip()
        if not name or seq is None:
            continue  # 跳过空行/汇总行

        item = {
            '序号': seq,
            '设备名称': name,
            '设备编号': str(vals[2] or '').strip(),
            '数量': str(vals[3] or '').strip(),
            '结构': str(vals[4] or '').strip(),
            '出图时间': vals[5],
            '采购': str(vals[6] or '').strip(),
            '机加件': str(vals[7] or '').strip(),
            '装配（硬件+电控）': str(vals[8] or '').strip(),
            '电气': str(vals[9] or '').strip(),
            '软件': str(vals[10] or '').strip(),
            '备注': str(vals[11] or '').strip(),
        }
        items.append(item)

    wb.close()
    return items


def get_project_title():
    """自动判断周次"""
    week_num = datetime.now().isocalendar()[1]
    return f"天勃新厂项目进度  |  第{week_num}周"


def generate_board(items, title=None, sort_by='progress', output_path=None):
    """生成进度看板"""

    # ── 解析数据 ──
    rows = []
    for item in items:
        if '只报价暂不生产' in (item.get('结构') or ''):
            continue  # 跳过暂不生产的

        progress, status = stage_to_progress(item)
        prio = parse_item_priority(item, progress, status)
        due = determine_due(item)

        rows.append({
            'name': item['设备名称'],
            'code': item['设备编号'],
            'qty': item['数量'],
            'priority': prio,
            'progress': progress,
            'status': status,
            'due': due,
        })

    # ── 排序 ──
    if sort_by == 'priority':
        prio_order = {'P0': 0, 'P1': 1, 'P2': 2}
        rows.sort(key=lambda r: (prio_order.get(r['priority'], 3), -r['progress']))
    elif sort_by == 'due':
        rows.sort(key=lambda r: r['due'] or 'Z')
    elif sort_by == 'name':
        rows.sort(key=lambda r: r['name'])
    else:  # progress
        rows.sort(key=lambda r: r['progress'])

    # ── 图表设置 ──
    font_name = find_chinese_font()
    if font_name:
        plt.rcParams['font.family'] = font_name
    plt.rcParams['axes.unicode_minus'] = False

    n_rows = len(rows)
    fig_height = max(4, 2.8 + n_rows * 0.95 + 0.8)
    fig_width = 16

    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    ax.set_xlim(0, 16)
    ax.set_ylim(0, fig_height)
    ax.axis('off')

    fig.patch.set_facecolor(BG_PAGE)
    ax.set_facecolor(BG_PAGE)

    # ── 标题 ──
    default_title = get_project_title()
    title_text = title or default_title
    ax.text(0.5, fig_height - 0.2, title_text,
            fontsize=20, fontweight='bold', color=TEXT_TITLE,
            ha='left', va='top')

    # 统计
    done_count = sum(1 for r in rows if r['progress'] >= 100)
    p0_count = sum(1 for r in rows if r['priority'] == 'P0')
    ax.text(0.5, fig_height - 0.7,
            f"生成: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
            f"共 {n_rows} 项  |  完成 {done_count} 项  |  紧急 {p0_count} 项",
            fontsize=10, color=TEXT_SUBTITLE, ha='left', va='top')

    # ── 列定义 ──
    cols = [
        ('优先级', 0.3, 1.0),
        ('设备名称', 1.5, 2.5),
        ('设备编号', 4.2, 2.0),
        ('数量', 6.4, 0.8),
        ('进度', 7.4, 2.5),
        ('状态', 10.1, 1.8),
        ('预计交期', 12.1, 1.8),
    ]

    header_y = fig_height - 1.2
    row_height = 0.95
    start_y = header_y - row_height

    # 表头
    header_rect = mpatches.FancyBboxPatch(
        (0.15, header_y - 0.25), 15.7, row_height,
        boxstyle="round,pad=0.1", facecolor=BG_HEADER, edgecolor='none'
    )
    ax.add_patch(header_rect)

    for col_name, x_start, width in cols:
        ax.text(x_start + width / 2, header_y + 0.15, col_name,
                fontsize=12, fontweight='bold', color=TEXT_HEADER,
                ha='center', va='center')

    # ── 数据行 ──
    current_y = start_y
    for idx, row in enumerate(rows):
        prio = row['priority']
        p_color = PRIORITY_COLORS.get(prio, PRIORITY_COLORS['P2'])
        progress = row['progress']

        # 行背景
        bg_color = ROW_EVEN if idx % 2 == 0 else ROW_ODD
        row_rect = mpatches.FancyBboxPatch(
            (0.15, current_y - row_height + 0.1), 15.7, row_height - 0.1,
            boxstyle="round,pad=0.02", facecolor=bg_color, edgecolor=ROW_BORDER,
            linewidth=0.5
        )
        ax.add_patch(row_rect)

        tag_y = current_y - row_height / 2 + 0.08

        # 优先级标签
        px, pw = cols[0][1], cols[0][2]
        tag_rect = mpatches.FancyBboxPatch(
            (px, tag_y - 0.25), pw, 0.5,
            boxstyle="round,pad=0.1", facecolor=p_color['tag'], edgecolor='none'
        )
        ax.add_patch(tag_rect)
        ax.text(px + pw / 2, tag_y + 0.02, prio,
                fontsize=11, fontweight='bold', color='white',
                ha='center', va='center')

        # 设备名称
        ax.text(cols[1][1] + 0.15, tag_y + 0.02, row['name'],
                fontsize=11, fontweight='bold', color='#212529',
                ha='left', va='center')

        # 设备编号（小字）
        ax.text(cols[2][1] + 0.1, tag_y + 0.02, row['code'],
                fontsize=8.5, color='#6C757D',
                ha='left', va='center')

        # 数量
        ax.text(cols[3][1] + cols[3][2] / 2, tag_y + 0.02, row['qty'],
                fontsize=10, color='#495057',
                ha='center', va='center')

        # 进度条
        bar_x = cols[4][1]
        bar_y = tag_y - 0.18
        bar_w = cols[4][2]
        bar_h = 0.36

        # 背景
        prog_bg = mpatches.FancyBboxPatch(
            (bar_x, bar_y), bar_w, bar_h,
            boxstyle="round,pad=0.03", facecolor=PROGRESS_BG, edgecolor='none'
        )
        ax.add_patch(prog_bg)

        # 填充
        if progress > 0:
            if progress >= 90:
                pc = PROGRESS_COLORS['high']
            elif progress >= 30:
                pc = PROGRESS_COLORS['mid']
            else:
                pc = PROGRESS_COLORS['low']

            fill_w = bar_w * (progress / 100)
            prog_fill = mpatches.FancyBboxPatch(
                (bar_x, bar_y), fill_w, bar_h,
                boxstyle="round,pad=0.03", facecolor=pc,
                edgecolor='none', alpha=0.85
            )
            ax.add_patch(prog_fill)

        # 进度百分比
        text_color = '#FFFFFF' if progress >= 50 else '#495057'
        ax.text(bar_x + bar_w / 2, tag_y + 0.02, f'{progress}%',
                fontsize=10, fontweight='bold', color=text_color,
                ha='center', va='center')

        # 状态文字
        ax.text(cols[5][1] + 0.1, tag_y + 0.02, row['status'],
                fontsize=10, color='#212529',
                ha='left', va='center')

        # 预计交期
        ax.text(cols[6][1] + cols[6][2] / 2, tag_y + 0.02, row['due'],
                fontsize=9, color='#6C757D',
                ha='center', va='center')

        current_y -= row_height

    # ── 图例 ──
    legend_y = current_y - 0.3
    ax.text(0.3, legend_y, '优先级：',
            fontsize=9, color='#6C757D', ha='left', va='center')

    legends = [
        ('P0 - 紧急(采购滞后)', PRIORITY_COLORS['P0']['tag']),
        ('P1 - 重要(推进中)', PRIORITY_COLORS['P1']['tag']),
        ('P2 - 常规/已完成', PRIORITY_COLORS['P2']['tag']),
    ]
    lx = 1.8
    for label, color in legends:
        rect = mpatches.FancyBboxPatch(
            (lx, legend_y - 0.18), 0.4, 0.36,
            boxstyle="round,pad=0.02", facecolor=color, edgecolor='none'
        )
        ax.add_patch(rect)
        ax.text(lx + 0.5, legend_y + 0.02, label,
                fontsize=8, color='#6C757D', ha='left', va='center')
        lx += 3.5

    # 进度条图例
    ax.text(10.5, legend_y, '进度颜色：',
            fontsize=9, color='#6C757D', ha='left', va='center')
    prog_legends = [
        ('≥90%', PROGRESS_COLORS['high']),
        ('30-89%', PROGRESS_COLORS['mid']),
        ('<30%', PROGRESS_COLORS['low']),
    ]
    plx = 12.5
    for label, color in prog_legends:
        rect = mpatches.FancyBboxPatch(
            (plx, legend_y - 0.18), 0.4, 0.36,
            boxstyle="round,pad=0.02", facecolor=color, edgecolor='none'
        )
        ax.add_patch(rect)
        ax.text(plx + 0.5, legend_y + 0.02, label,
                fontsize=8, color='#6C757D', ha='left', va='center')
        plx += 1.8

    # ── 输出 ──
    script_dir = os.path.dirname(os.path.abspath(__file__))
    if output_path is None:
        output_path = os.path.join(script_dir, '天勃项目进度看板.png')

    plt.tight_layout()
    plt.savefig(output_path, dpi=200, bbox_inches='tight',
                facecolor=fig.get_facecolor(), edgecolor='none')
    plt.close(fig)

    return output_path


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(script_dir, '天勃项目进度跟踪表.xlsx123.xlsx')

    # 参数
    sort_by = 'progress'
    title = None
    i = 0
    while i < len(sys.argv):
        arg = sys.argv[i]
        if arg == '--sort' and i + 1 < len(sys.argv):
            sort_by = sys.argv[i + 1]
            i += 1
        elif arg == '--title' and i + 1 < len(sys.argv):
            title = sys.argv[i + 1]
            i += 1
        i += 1

    if not os.path.exists(xlsx_path):
        # 尝试找实际文件
        import glob
        found = glob.glob(os.path.join(script_dir, '天勃*.xlsx*'))
        if found:
            xlsx_path = found[0]
        else:
            print(f"❌ 找不到进度表xlsx文件")
            sys.exit(1)

    print(f"📂 读取: {xlsx_path}")
    items = load_xlsx_data(xlsx_path)

    # 去重（相同设备编号只留一条）
    seen = {}
    for item in items:
        code = item['设备编号']
        if code and code not in seen:
            seen[code] = item
        elif not code:
            # 无编号的用名称去重
            if item['设备名称'] not in seen:
                seen[item['设备名称']] = item

    items = list(seen.values())
    print(f"📊 共 {len(items)} 个设备项目")

    output = generate_board(items, title=title, sort_by=sort_by)
    print(f"✅ 看板图已生成: {output}")

    # 同步到 workspace 根目录
    workspace_root = os.path.abspath(os.path.join(script_dir, '..', '..'))
    dest = os.path.join(workspace_root, '天勃项目进度看板.png')
    import shutil
    shutil.copy2(output, dest)
    print(f"📋 已同步到: {dest}")

    # 打印摘要
    print(f"\n📋 进度摘要:")
    print(f"{'设备名称':15s} {'进度':>6s} {'优先级'} {'状态':12s}")
    print('-' * 50)
    for item in items:
        progress, status = stage_to_progress(item)
        if progress == 0 and status == '暂不进行':
            continue
        prio = parse_item_priority(item, progress, status)
        name = item['设备名称'][:12]
        print(f"{name:15s} {progress:4d}%  {prio:4s}  {status:12s}")


if __name__ == '__main__':
    main()
